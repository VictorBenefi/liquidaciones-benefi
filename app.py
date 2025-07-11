import streamlit as st
import pandas as pd
from io import BytesIO
from openpyxl.styles import Font, Alignment
from datetime import datetime
import os

# --- LOGIN ---
USUARIOS_AUTORIZADOS = {
    "admin": "clave123",
    "victor": "benefi2024",
}

if "logueado" not in st.session_state:
    st.session_state.logueado = False
    st.session_state.usuario = ""

if not st.session_state.logueado:
    st.title("🔒 Ingreso a BENEFI")

    with st.form("login_form"):
        usuario = st.text_input("Usuario")
        clave = st.text_input("Contraseña", type="password")
        enviar = st.form_submit_button("Ingresar")

    if enviar:
        if usuario in USUARIOS_AUTORIZADOS and USUARIOS_AUTORIZADOS[usuario] == clave:
            st.session_state.logueado = True
            st.session_state.usuario = usuario
            st.success("✅ Ingreso exitoso. Esperá un segundo...")
        else:
            st.error("Usuario o contraseña incorrectos")

    if not st.session_state.logueado:
        st.stop()

# --- APP PRINCIPAL ---
st.success(f"Bienvenido {st.session_state.usuario} 👋")
st.title("💰 Sistema de Liquidaciones con IVA - Benefi")

if st.button("Cerrar sesión 🔒"):
    st.session_state.logueado = False
    st.session_state.usuario = ""
    st.info("Sesión cerrada. Recargá la página para volver al login.")
    st.stop()

archivo = st.file_uploader("📁 Subí el archivo Excel", type=["xlsx"])

if archivo:
    df = pd.read_excel(archivo)

    columnas_necesarias = {"red", "Total_Ventas", "Cantidad_Ventas", "Costo_Amin", "Costo_Tr"}

    if columnas_necesarias.issubset(df.columns):
        df["Costo_Admin"] = df["Total_Ventas"] * df["Costo_Amin"]
        df["Costo_Transaccion"] = df["Cantidad_Ventas"] * df["Costo_Tr"]
        df["Subtotal"] = df["Costo_Admin"] + df["Costo_Transaccion"]
        df["IVA_21%"] = df["Subtotal"] * 0.21
        df["Total_Cobrar"] = df["Subtotal"] + df["IVA_21%"]

        st.subheader("📊 Resultado de la liquidación")
        st.dataframe(df)

        hoy = datetime.today()
        nombre_mes = hoy.strftime('%B').capitalize()
        fecha_str = hoy.strftime('%d de %B de %Y').capitalize()

        salida = BytesIO()
        with pd.ExcelWriter(salida, engine="openpyxl") as writer:
            df.to_excel(writer, index=False, sheet_name="Liquidación", startrow=3)
            hoja = writer.sheets["Liquidación"]

            hoja.merge_cells("A1:F1")
            celda1 = hoja["A1"]
            celda1.value = "BENEFI - LIQUIDACIÓN MENSUAL"
            celda1.font = Font(size=14, bold=True)
            celda1.alignment = Alignment(horizontal="center")

            hoja.merge_cells("A2:F2")
            celda2 = hoja["A2"]
            celda2.value = f"Corresponde a {nombre_mes.upper()} {hoy.year} - Generado el {fecha_str}"
            celda2.font = Font(size=11, italic=True)
            celda2.alignment = Alignment(horizontal="center")

            columnas_monedas = ["Costo_Admin", "Costo_Transaccion", "Subtotal", "IVA_21%", "Total_Cobrar"]
            for col in hoja.iter_cols(min_row=4, max_row=hoja.max_row):
                if col[0].value in columnas_monedas:
                    for celda in col[1:]:
                        celda.number_format = '# ##0,00'

        st.download_button(
            label="📥 Descargar Excel con encabezado y formato",
            data=salida.getvalue(),
            file_name="Cobrar_liquidacion_junio.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

    else:
        st.error("❗ El archivo debe tener las columnas: red, Total_Ventas, Cantidad_Ventas, Costo_Amin, Costo_Tr")

st.markdown("---")
st.header("📁 Historial de Liquidaciones")

HISTORIAL_DIR = "historial"
os.makedirs(HISTORIAL_DIR, exist_ok=True)

archivos = sorted([f for f in os.listdir(HISTORIAL_DIR) if f.endswith(".xlsx")], reverse=True)

if archivos:
    for archivo in archivos:
        col1, col2, col3 = st.columns([4, 2, 1])
        with col1:
            st.write(f"📄 {archivo}")
        with col2:
            with open(os.path.join(HISTORIAL_DIR, archivo), "rb") as f:
                st.download_button(
                    label="📥 Descargar",
                    data=f,
                    file_name=archivo,
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    key=f"descarga_{archivo}"
                )
        with col3:
            if st.button("🗑️ Eliminar", key=f"eliminar_{archivo}"):
                os.remove(os.path.join(HISTORIAL_DIR, archivo))
                st.success(f"Archivo eliminado: {archivo}")
                st.experimental_rerun()
else:
    st.info("Todavía no se han generado liquidaciones.")
